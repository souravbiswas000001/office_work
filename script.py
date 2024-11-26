import os
import shutil
from openpyxl import load_workbook

# Define file paths
reference_file = "AppReferenceIDs.xlsx"  # Path to the reference data file
template_file = "DataIntakeform.xlsx"  # Path to the template file
output_dir = "/Users/souravbiswas/PycharmProjects/script_python/output"  # Directory for output files

# Create output directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# Load the reference data
wb_ref = load_workbook(reference_file)
sheet_ref = wb_ref.active

# Read reference data (assuming headers are in the first row)
for row in sheet_ref.iter_rows(min_row=2, values_only=True):  # Skip header
    app_id = row[1]  # Second column
    ref_value = row[2]  # Third column

    if not app_id or not ref_value:
        print("Skipping row due to missing data:", row)
        continue

    # Generate output filename using AppID and reference data
    output_filename = f"{app_id}-{ref_value}"
    print(f"Processing: {output_filename}")

    # Generate the new file name
    new_filename = os.path.join(output_dir, f"{output_filename}.xlsx")

    # Copy the template file
    shutil.copy(template_file, new_filename)

    # Open the copied file and populate the defined cell
    wb_new = load_workbook(new_filename)
    sheet_new = wb_new.active

    defined_cell = "A1"
    sheet_new[defined_cell] = app_id
    print(f"Populated cell {defined_cell} with {app_id} in {new_filename}")

    # Save the updated file
    wb_new.save(new_filename)
    wb_new.close()

# Close the reference workbook
wb_ref.close()

print(f"Process completed. Files saved in '{output_dir}'.")